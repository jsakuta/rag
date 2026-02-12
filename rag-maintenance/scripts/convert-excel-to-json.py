"""
Excel -> JSON 変換スクリプト
rag-gemini の実データ（Excel）を Cosmos DB 投入用 JSON に変換する。

使用法:
  pip install openpyxl
  python scripts/convert-excel-to-json.py

出力:
  scripts/data/scenarios.json
  scripts/data/faqs.json
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

# --- 設定 ---
RAG_GEMINI_BASE = Path(r"C:\VSCode\rag\rag-gemini\data\source")
OUTPUT_DIR = Path(__file__).parent / "data"

SCENARIO_FILES = [
    {
        "path": RAG_GEMINI_BASE / "scenarios" / "revisions" / "rev03smile_シナリオデータ_20260203.xlsx",
        "categoryId": "smile",
        "categoryName": "スマイル",
    },
    {
        "path": RAG_GEMINI_BASE / "scenarios" / "revisions" / "rev03souzoku_シナリオデータ_20260203.xlsx",
        "categoryId": "souzoku",
        "categoryName": "相続",
    },
    {
        "path": RAG_GEMINI_BASE / "scenarios" / "revisions" / "rev03naibujimu_シナリオデータ_20260203.xlsx",
        "categoryId": "naibujimu",
        "categoryName": "内部事務",
    },
    {
        "path": RAG_GEMINI_BASE / "scenarios" / "revisions" / "rev03torikaku_シナリオデータ_20260203.xlsx",
        "categoryId": "torikaku",
        "categoryName": "取引時確認",
    },
]

FAQ_FILES = [
    {
        "path": RAG_GEMINI_BASE / "faq" / "latest" / "スマイル_履歴データ_20250205.xlsx",
        "categoryId": "smile",
        "categoryName": "スマイル",
        "columns": {"title": "問い合わせ", "content": "回答", "supplement": "補足回答", "tags": None},
    },
    {
        "path": RAG_GEMINI_BASE / "faq" / "latest" / "総則_履歴データ_20250829.xlsx",
        "categoryId": "sousoku",
        "categoryName": "総則",
        "columns": {"title": "質問", "content": "回答", "supplement": None, "tags": "タグ付け"},
    },
    {
        "path": RAG_GEMINI_BASE / "faq" / "latest" / "預金_履歴データ_20250830.xlsx",
        "categoryId": "yokin",
        "categoryName": "預金",
        "columns": {"title": "質問", "content": "回答", "supplement": None, "tags": "タグ付け"},
    },
]

UPDATED_AT = "2026-02-12T00:00:00+09:00"


def normalize_text(text: str) -> str:
    """Excel特有のキャリッジリターンを正規化する。"""
    if not text:
        return ""
    text = text.replace("_x000D_\n", "\n")
    text = text.replace("_x000D_", "")
    text = text.replace("\r\n", "\n")
    text = text.replace("\r", "\n")
    return text.strip()


def cell_value(cell) -> str:
    """セル値を文字列として取得。None/空はから文字列を返す。"""
    v = cell
    if v is None:
        return ""
    s = str(v).strip()
    return normalize_text(s)


# ==============================
# シナリオ変換
# ==============================

def get_lv_columns(ws):
    """ヘッダー行からLv列のインデックスを取得する。"""
    headers = []
    for cell in ws[1]:
        val = cell_value(cell.value)
        if val.startswith("Lv"):
            headers.append(cell.column - 1)  # 0-indexed
    return headers


def parse_row_levels(row_values, lv_indices):
    """行のLv値をリストとして返す。"""
    levels = []
    for idx in lv_indices:
        if idx < len(row_values):
            levels.append(cell_value(row_values[idx]))
        else:
            levels.append("")
    return levels


def get_deepest_level(levels):
    """最右の非空Lvのインデックスを返す（0始まり）。全て空なら-1。"""
    for i in range(len(levels) - 1, -1, -1):
        if levels[i]:
            return i
    return -1


def get_path_up_to(levels, depth):
    """depth までのパスをタプルとして返す。"""
    return tuple(levels[i] for i in range(depth + 1) if i < len(levels))


def convert_scenarios():
    """シナリオExcelファイルをJSON形式に変換する。"""
    all_docs = []
    total_skipped = 0

    for file_config in SCENARIO_FILES:
        filepath = file_config["path"]
        category_id = file_config["categoryId"]
        category_name = file_config["categoryName"]

        if not filepath.exists():
            print(f"  [WARN] ファイルが見つかりません: {filepath}")
            continue

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # ヘッダーからLv列を特定
        lv_indices = get_lv_columns(ws)
        if not lv_indices:
            print(f"  [WARN] Lv列が見つかりません: {filepath.name}")
            wb.close()
            continue

        # 全行のレベル情報を先読み（isFinalAnswer判定用）
        all_rows_raw = list(ws.iter_rows(min_row=2, values_only=True))
        all_levels = []
        for row_vals in all_rows_raw:
            levels = parse_row_levels(row_vals, lv_indices)
            all_levels.append(levels)

        seq = 0
        skipped = 0

        for row_idx, levels in enumerate(all_levels):
            depth = get_deepest_level(levels)
            if depth < 0:
                skipped += 1
                continue  # 空行スキップ

            # 回答（最右非空セル）
            answer = levels[depth]
            if not answer:
                skipped += 1
                continue

            # 質問（回答の1つ左）
            question = levels[depth - 1] if depth >= 1 else ""

            # 分類（質問より左の全レベル）
            hierarchy_parts = [levels[i] for i in range(max(0, depth - 1)) if levels[i]]

            # path / title
            path_parts = hierarchy_parts + ([question] if question else [])
            path = "/" + "/".join(path_parts) if path_parts else "/"
            title = path

            # combinedContent
            hierarchy_str = " > ".join(hierarchy_parts) if hierarchy_parts else ""
            if hierarchy_str and question:
                combined = f"分類: {hierarchy_str} | 質問: {question} | 回答: {answer}"
            elif question:
                combined = f"質問: {question} | 回答: {answer}"
            else:
                combined = f"回答: {answer}"

            # keywords
            keywords = [lv for lv in hierarchy_parts if lv]
            if question and question not in keywords:
                keywords.append(question)

            # isFinalAnswer 判定
            is_final = True
            current_path = get_path_up_to(levels, depth - 1) if depth >= 1 else ()
            for next_idx in range(row_idx + 1, len(all_levels)):
                next_depth = get_deepest_level(all_levels[next_idx])
                if next_depth < 0:
                    continue
                next_path = get_path_up_to(all_levels[next_idx], depth - 1) if depth >= 1 else ()
                if next_path == current_path and next_depth > depth:
                    is_final = False
                    break
                if next_path != current_path:
                    break

            seq += 1
            doc = {
                "id": f"scenario-{category_id}-{seq:04d}",
                "dataType": "scenario",
                "categoryId": category_id,
                "categoryName": category_name,
                "title": title,
                "content": answer,
                "combinedContent": combined,
                "keywords": keywords,
                "updatedAt": UPDATED_AT,
                "isDeleted": False,
                "path": path,
                "order": row_idx + 1,
                "isFinalAnswer": is_final,
            }
            all_docs.append(doc)

        wb.close()
        total_skipped += skipped
        print(f"  {filepath.name}: {seq}件変換 ({skipped}件スキップ)")

    print(f"  シナリオ合計: {len(all_docs)}件 ({total_skipped}件スキップ)")
    return all_docs


# ==============================
# FAQ 変換
# ==============================

def find_column_index(ws, column_name):
    """ヘッダー行から列名のインデックスを返す（1始まり）。見つからなければNone。"""
    for cell in ws[1]:
        if cell_value(cell.value) == column_name:
            return cell.column
    return None


def parse_tags_to_keywords(tag_str: str) -> list:
    """タグ文字列からキーワードを抽出する。
    例: "Lv0:預金 Lv1:口座開設" -> ["預金", "口座開設"]
    """
    if not tag_str:
        return []
    keywords = []
    for match in re.finditer(r"Lv\d+:\s*([^\s]+)", tag_str):
        kw = match.group(1).strip()
        if kw:
            keywords.append(kw)
    return keywords


def convert_faqs():
    """FAQ ExcelファイルをJSON形式に変換する。"""
    all_docs = []
    total_skipped = 0

    for file_config in FAQ_FILES:
        filepath = file_config["path"]
        category_id = file_config["categoryId"]
        category_name = file_config["categoryName"]
        col_config = file_config["columns"]

        if not filepath.exists():
            print(f"  [WARN] ファイルが見つかりません: {filepath}")
            continue

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # 列インデックスを検出
        title_col = find_column_index(ws, col_config["title"])
        content_col = find_column_index(ws, col_config["content"])
        supplement_col = find_column_index(ws, col_config["supplement"]) if col_config["supplement"] else None
        tags_col = find_column_index(ws, col_config["tags"]) if col_config["tags"] else None

        if not title_col or not content_col:
            print(f"  [WARN] 必須列が見つかりません: {filepath.name} (title={col_config['title']}, content={col_config['content']})")
            wb.close()
            continue

        seq = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=False):
            title_val = cell_value(row[title_col - 1].value)
            content_val = cell_value(row[content_col - 1].value)

            # 空行スキップ
            if not title_val and not content_val:
                skipped += 1
                continue

            # テストデータ行スキップ（"テスト" のみの行）
            if title_val == "テスト" or content_val == "テスト":
                skipped += 1
                continue

            # content必須チェック
            if not content_val:
                skipped += 1
                continue

            # titleが空の場合もスキップ
            if not title_val:
                skipped += 1
                continue

            # 補足回答の連結
            supplement_val = ""
            if supplement_col and (supplement_col - 1) < len(row):
                supplement_val = cell_value(row[supplement_col - 1].value)
            if supplement_val:
                content = content_val + "\n\n" + supplement_val
            else:
                content = content_val

            # タグ
            tags_val = None
            if tags_col and (tags_col - 1) < len(row):
                tags_val = cell_value(row[tags_col - 1].value) or None

            # keywords
            keywords = parse_tags_to_keywords(tags_val) if tags_val else []

            # combinedContent
            combined = f"質問: {title_val} | 回答: {content}"

            seq += 1
            doc = {
                "id": f"faq-{category_id}-{seq:05d}",
                "dataType": "faq",
                "categoryId": category_id,
                "categoryName": category_name,
                "title": title_val,
                "content": content,
                "combinedContent": combined,
                "keywords": keywords,
                "updatedAt": UPDATED_AT,
                "isDeleted": False,
                "tags": tags_val,
            }
            all_docs.append(doc)

        wb.close()
        total_skipped += skipped
        print(f"  {filepath.name}: {seq}件変換 ({skipped}件スキップ)")

    print(f"  FAQ合計: {len(all_docs)}件 ({total_skipped}件スキップ)")
    return all_docs


# ==============================
# バリデーション
# ==============================

def validate(docs, data_type):
    """生成されたドキュメントのバリデーション。"""
    errors = 0
    ids = set()

    for doc in docs:
        # 必須フィールドチェック
        for field in ["id", "categoryId", "title", "content", "combinedContent"]:
            if not doc.get(field):
                print(f"  [ERROR] {doc['id']}: {field} が空です")
                errors += 1

        # ID重複チェック
        if doc["id"] in ids:
            print(f"  [ERROR] ID重複: {doc['id']}")
            errors += 1
        ids.add(doc["id"])

    if errors == 0:
        print(f"  {data_type} バリデーション OK ({len(docs)}件)")
    else:
        print(f"  {data_type} バリデーション NG ({errors}件のエラー)")

    return errors


# ==============================
# メイン
# ==============================

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("Excel -> JSON 変換スクリプト")
    print("=" * 60)

    # シナリオ変換
    print("\n--- シナリオ変換 ---")
    scenarios = convert_scenarios()

    # FAQ変換
    print("\n--- FAQ変換 ---")
    faqs = convert_faqs()

    # バリデーション
    print("\n--- バリデーション ---")
    s_errors = validate(scenarios, "scenarios")
    f_errors = validate(faqs, "faqs")

    if s_errors + f_errors > 0:
        print(f"\n[ERROR] バリデーションエラーが {s_errors + f_errors}件あります。出力を中断します。")
        sys.exit(1)

    # JSON出力
    scenarios_path = OUTPUT_DIR / "scenarios.json"
    faqs_path = OUTPUT_DIR / "faqs.json"

    with open(scenarios_path, "w", encoding="utf-8") as f:
        json.dump(scenarios, f, ensure_ascii=False, indent=2)
    print(f"\n出力: {scenarios_path} ({len(scenarios)}件)")

    with open(faqs_path, "w", encoding="utf-8") as f:
        json.dump(faqs, f, ensure_ascii=False, indent=2)
    print(f"出力: {faqs_path} ({len(faqs)}件)")

    # サマリー
    print("\n" + "=" * 60)
    print(f"シナリオ: {len(scenarios)}件")
    for cfg in SCENARIO_FILES:
        count = sum(1 for d in scenarios if d["categoryId"] == cfg["categoryId"])
        print(f"  {cfg['categoryName']}: {count}件")
    print(f"FAQ: {len(faqs)}件")
    for cfg in FAQ_FILES:
        count = sum(1 for d in faqs if d["categoryId"] == cfg["categoryId"])
        print(f"  {cfg['categoryName']}: {count}件")
    print("=" * 60)


if __name__ == "__main__":
    main()
